"""
Excel 读取工具 — 从原 huili/惠理基金/src/etl_excel/extract.py 迁移
使用 openpyxl read_only 模式提升大文件读取性能。
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook


def list_sheet_names(input_path: Path) -> List[str]:
    """返回 Excel 中所有 Sheet 名称列表"""
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def read_sheet_rows(input_path: Path, sheet_name: str) -> List[List[object]]:
    """读取指定 Sheet 的所有行数据，返回二维列表"""
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        rows: List[List[object]] = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append(list(row))
        return rows
    finally:
        workbook.close()


def read_workbook_rows(input_path: Path) -> Dict[str, List[List[object]]]:
    """读取整个 Excel 文件，返回 {sheet_name: rows} 字典"""
    return {name: read_sheet_rows(input_path, name) for name in list_sheet_names(input_path)}
