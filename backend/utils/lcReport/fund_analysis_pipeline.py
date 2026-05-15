"""
ETL Pipeline — 从原 huili/惠理基金/src/etl_excel/pipeline.py 迁移并适配
阶段一：Excel → CSV（内存中的结构化 DataFrame）
读取每张 Sheet → 解析表头 → 解析数据行 → 清洗 → 校验
"""
from __future__ import annotations

import json
import uuid
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd

from .cleaner import apply_type_rules, standardize_placeholders, summarize_conversion_errors
from .config import EtlConfig, PipelineArgs
from .errors import EtlErrorContext, EtlRuleError
from .extract import list_sheet_names, read_sheet_rows
from .header_parser import parse_header
from .meta_extract import parse_sheet_metadata
from .row_parser import parse_rows
from .validator import summarize_issues, validate_sheet


def _ensure_output_dir(output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)


def _drop_unnamed_columns(frame: pd.DataFrame) -> pd.DataFrame:
    unnamed_cols = [col for col in frame.columns if str(col).startswith("Unnamed_")]
    if unnamed_cols:
        frame = frame.drop(columns=unnamed_cols)
    return frame


def _parse_single_sheet(
    input_path: Path, sheet_name: str, config: EtlConfig, etl_run_id: str
) -> Dict[str, object]:
    """解析单张 Sheet，返回包含 frame / header / metadata / issues 的结果字典"""
    rows = read_sheet_rows(input_path, sheet_name)
    if not rows:
        raise EtlRuleError(
            "[E3001] empty worksheet",
            EtlErrorContext(sheet_name=sheet_name, row_idx=None, col_idx=None, rule_id="sheet_non_empty"),
        )

    header = parse_header(rows, config, sheet_name)
    data_rows = rows[header.data_start_row_idx:]
    parsed_rows = parse_rows(
        data_rows=data_rows,
        flat_columns=header.flat_columns,
        sheet_name=sheet_name,
        start_row_idx=header.data_start_row_idx,
        config=config,
    )
    frame = pd.DataFrame([r.payload for r in parsed_rows])
    if frame.empty:
        frame = pd.DataFrame(columns=[*header.flat_columns, "group_category", "row_source_sheet", "row_number"])

    frame = _drop_unnamed_columns(frame)
    frame = standardize_placeholders(frame, config)
    frame, conversion_errors = apply_type_rules(frame, config)
    metadata = parse_sheet_metadata(rows, sheet_name, input_path)
    frame["report_set"] = metadata["report_set"]
    frame["etl_run_id"] = etl_run_id
    issues = validate_sheet(frame, sheet_name, config)

    return {
        "sheet_name": sheet_name,
        "frame": frame,
        "header": header,
        "conversion_errors": conversion_errors,
        "issues": issues,
        "metadata": metadata,
    }


def extract_qw_inception_dates(input_path: Path) -> Dict[str, Optional[str]]:
    """
    从 Quartile_weekly 文件的 inception 列头提取 fund_code → inception_date 映射。

    QW 文件的 HKSFC funds 等 sheet 中：
      Row 7: 'VPAF (inception)', 'VPHY (inception)' ...  (即 fund_code)
      Row 8: '4/2/1993', '9/3/2002' ...                 (即 inception_date)

    返回示例: {'VPAF': '1993-04-02', 'VPHY': '2002-09-03', ...}
    """
    import openpyxl
    from datetime import datetime

    wb = openpyxl.load_workbook(str(input_path), data_only=True)
    result: Dict[str, Optional[str]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_col = ws.max_column
        for j in range(1, max_col + 1):
            hdr = ws.cell(7, j).value
            if not hdr or '(inception)' not in str(hdr):
                continue
            fund_code = str(hdr).replace('(inception)', '').strip()
            if fund_code in result:
                continue  # 已经从其他 sheet 取到
            raw_date = ws.cell(8, j).value
            if raw_date is None:
                continue
            if hasattr(raw_date, 'strftime'):
                result[fund_code] = raw_date.strftime('%Y-%m-%d')
            else:
                # 字符串格式，如 '4/2/1993'
                raw_str = str(raw_date).strip()
                for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y'):
                    try:
                        result[fund_code] = datetime.strptime(raw_str, fmt).strftime('%Y-%m-%d')
                        break
                    except ValueError:
                        continue

    wb.close()
    return result


def run_fund_analysis_pipeline(
    input_path: Path,
    output_dir: Path | None = None,
    mode: str = "lenient",
) -> Dict[str, object]:
    """
    主入口：运行完整 ETL Pipeline（FundAnalysis 阶段一）。

    Args:
        input_path: Excel 文件路径
        output_dir: 中间产物输出目录（可选，不传则只在内存中处理）
        mode: "strict"（质量错误时抛出异常）或 "lenient"（继续执行）
        report_type: 报告类型，若是 Quartile_weekly 则按照 config.qw_target_sheets 过滤 sheet

    Returns:
        包含以下键的结果字典：
        - etl_run_id: 本次运行的 UUID
        - parsed_df: 合并后的 DataFrame（所有 Sheet 数据）
        - meta_records: 元数据字典列表
        - quality: 质量报告摘要
        - conversion: 类型转换错误摘要
    """
    args = PipelineArgs(input_path=input_path, output_dir=output_dir or input_path.parent, mode=mode)
    if not input_path.exists():
        raise FileNotFoundError(f"[E1001] input file not found: {input_path.as_posix()}")
    if mode not in {"strict", "lenient"}:
        mode = "lenient"

    config = EtlConfig(strict_mode=(mode == "strict"))
    if output_dir:
        _ensure_output_dir(output_dir)
    etl_run_id = str(uuid.uuid4())

    all_results: List[Dict[str, object]] = []
    all_sheets = list_sheet_names(input_path)
    # FundAnalysis 排除一些常见的说明类 sheet，其余都解析
    target_sheets = [s for s in all_sheets if not s.lower().startswith("readme") and not s.lower().startswith("instruction")]

    for sheet_name in target_sheets:
        result = _parse_single_sheet(input_path, sheet_name, config, etl_run_id)
        all_results.append(result)
        if output_dir:
            safe_name = sheet_name.replace("/", "_").replace("\\", "_").replace(" ", "_")
            result["frame"].to_csv(output_dir / f"parsed_{safe_name}.csv", index=False, encoding="utf-8-sig")

    all_frames = [r["frame"] for r in all_results]
    merged = pd.concat(all_frames, ignore_index=True) if all_frames else pd.DataFrame()

    if output_dir:
        merged.to_csv(output_dir / "parsed_all.csv", index=False, encoding="utf-8-sig")
        pd.DataFrame([r["metadata"] | {"etl_run_id": etl_run_id} for r in all_results]).to_csv(
            output_dir / "report_meta.csv", index=False, encoding="utf-8-sig"
        )

    all_conversion_errors = [e for result in all_results for e in result["conversion_errors"]]
    all_issues = [i for result in all_results for i in result["issues"]]
    quality_summary = summarize_issues(all_issues)
    conversion_summary = summarize_conversion_errors(all_conversion_errors, config.max_error_samples)

    if config.strict_mode and quality_summary["errors"] > 0:
        raise RuntimeError(
            f"[E3002] strict mode failed with {quality_summary['errors']} quality errors"
        )

    return {
        "etl_run_id": etl_run_id,
        "mode": mode,
        "sheet_count": len(all_results),
        "row_count_total": int(len(merged)),
        "parsed_df": merged,
        "meta_records": [r["metadata"] | {"etl_run_id": etl_run_id} for r in all_results],
        "quality": quality_summary,
        "conversion": conversion_summary,
        "column_lineage_by_sheet": {
            r["sheet_name"]: r["header"].column_lineage for r in all_results
        },
    }
