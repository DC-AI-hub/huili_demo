"""
表头解析器 — 从原 huili/惠理基金/src/etl_excel/header_parser.py 迁移
处理 Quartile_weekly Excel 多级合并表头，将其打平为一维列名列表。
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Dict, List, Sequence, Tuple

from openpyxl.utils.cell import get_column_letter

from .config import EtlConfig
from .errors import EtlErrorContext, EtlRuleError


@dataclass
class HeaderParseResult:
    header_anchor_row_idx: int       # "Group/Investment" 所在行索引
    header_start_row_idx: int        # 合并表头真实起始行索引
    data_start_row_idx: int          # 数据行起始行索引（anchor 下一行）
    flat_columns: List[str]          # 打平后的一维列名列表
    column_lineage: Dict[str, Dict[str, object]]  # 列溯源信息


import pandas as pd

def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    if pd.isna(value):
        return ""
    return str(value).replace("\n", " ").strip()


def locate_header_anchor(rows: Sequence[Sequence[object]], config: EtlConfig, sheet_name: str) -> int:
    """定位包含 'Group/Investment' 锚点文本的行索引"""
    for row_idx, row in enumerate(rows):
        for cell in row:
            if _normalize_text(cell) == config.header_anchor:
                return row_idx
    raise EtlRuleError(
        "[E2001] failed to locate header anchor",
        EtlErrorContext(sheet_name=sheet_name, row_idx=None, col_idx=None, rule_id="header_anchor"),
    )


def detect_header_start(
    rows: Sequence[Sequence[object]], anchor_row_idx: int, config: EtlConfig, sheet_name: str
) -> int:
    """从锚点行向上探测合并表头的真实起始行"""
    header_start = anchor_row_idx
    for probe in range(anchor_row_idx - 1, max(anchor_row_idx - config.max_header_depth, -1), -1):
        row = rows[probe] if probe < len(rows) else []
        non_empty = any(_normalize_text(v) for v in row)
        if non_empty:
            header_start = probe
        else:
            break
    if header_start > anchor_row_idx:
        raise EtlRuleError(
            "[E2002] invalid header start detected",
            EtlErrorContext(sheet_name=sheet_name, row_idx=header_start, col_idx=None, rule_id="header_start"),
        )
    return header_start


def _build_column_name(parts: List[str], col_idx: int, used: set) -> str:
    """将多级表头的各层文本拼接为唯一列名"""
    non_empty_parts = [p for p in parts if p]
    deduped_parts: List[str] = []
    for part in non_empty_parts:
        if not deduped_parts:
            deduped_parts.append(part)
            continue
        if deduped_parts[-1] == part:
            continue
        if part.startswith(f"{deduped_parts[-1]} "):
            deduped_parts[-1] = part
            continue
        deduped_parts.append(part)
    base = "_".join(deduped_parts) if deduped_parts else f"Unnamed_{col_idx}"
    base = re.sub(r"\s+", " ", base).strip()
    if not base:
        base = f"Unnamed_{col_idx}"
    candidate = base
    if candidate in used:
        suffix = get_column_letter(col_idx + 1)
        candidate = f"{base}__{suffix}"
        counter = 1
        while candidate in used:
            candidate = f"{base}__{suffix}_{counter}"
            counter += 1
    used.add(candidate)
    return candidate


def flatten_header(
    rows: Sequence[Sequence[object]], header_start: int, anchor_row: int, sheet_name: str
) -> Tuple[List[str], Dict[str, Dict[str, object]]]:
    """将多级合并表头打平为一维列名列表，同时记录列溯源信息"""
    header_rows = rows[header_start: anchor_row + 1]
    width = max(len(row) for row in header_rows) if header_rows else 0
    if width == 0:
        raise EtlRuleError(
            "[E2003] empty header rows",
            EtlErrorContext(sheet_name=sheet_name, row_idx=header_start, col_idx=None, rule_id="header_flatten"),
        )

    flat_columns: List[str] = []
    column_lineage: Dict[str, Dict[str, object]] = {}
    used: set = set()
    normalized_grid: List[List[str]] = []

    for row_idx, row in enumerate(header_rows):
        normalized_row: List[str] = []
        for col_idx in range(width):
            cell_value = row[col_idx] if col_idx < len(row) else None
            normalized_row.append(_normalize_text(cell_value))
        # 非锚点行（上层合并表头）：向右填充空格（模拟合并单元格效果）
        if row_idx < len(header_rows) - 1:
            for col_idx in range(1, width):
                if not normalized_row[col_idx]:
                    normalized_row[col_idx] = normalized_row[col_idx - 1]
        normalized_grid.append(normalized_row)

    for col_idx in range(width):
        pieces = []
        for row in normalized_grid:
            text = row[col_idx]
            if text:
                pieces.append(text)
        col_name = _build_column_name(pieces, col_idx, used)
        col_letter = get_column_letter(col_idx + 1)
        flat_columns.append(col_name)
        column_lineage[col_name] = {
            "column_letter": col_letter,
            "source_rows": [header_start + 1, anchor_row + 1],
            "source_cells": [f"{col_letter}{header_start + 1}", f"{col_letter}{anchor_row + 1}"],
            "parts": pieces,
        }
    return flat_columns, column_lineage


def parse_header(rows: Sequence[Sequence[object]], config: EtlConfig, sheet_name: str) -> HeaderParseResult:
    """主入口：解析完整表头，返回打平列名和数据起始行信息"""
    anchor_row_idx = locate_header_anchor(rows, config, sheet_name)
    header_start_row_idx = detect_header_start(rows, anchor_row_idx, config, sheet_name)
    flat_columns, column_lineage = flatten_header(rows, header_start_row_idx, anchor_row_idx, sheet_name)
    return HeaderParseResult(
        header_anchor_row_idx=anchor_row_idx,
        header_start_row_idx=header_start_row_idx,
        data_start_row_idx=anchor_row_idx + 1,
        flat_columns=flat_columns,
        column_lineage=column_lineage,
    )
